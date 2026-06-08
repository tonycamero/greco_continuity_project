#!/usr/bin/env python3
"""
Create a reviewable Excel workbook from the hydrated relationship CSV.
"""

from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data" / "greco_relationships_hydrated.csv"
XLSX_PATH = ROOT / "data" / "greco_relationships_hydrated.xlsx"


def read_rows() -> tuple[list[str], list[dict[str, str]]]:
    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def set_widths(ws, headers: list[str], rows: list[dict[str, str]]) -> None:
    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in rows[:200]:
            max_len = max(max_len, len(str(row.get(header, ""))))
        if header in {"strategic_fit", "current_hook", "first_move", "risk_note", "source_urls", "notes"}:
            width = min(max(max_len * 0.8, 24), 58)
        elif header.endswith("_url") or header in {"website", "contact_url", "source_urls"}:
            width = min(max(max_len * 0.7, 18), 44)
        else:
            width = min(max(max_len * 0.9, 10), 28)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width


def build() -> None:
    headers, rows = read_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Relationships"

    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    set_widths(ws, headers, rows)

    table = Table(displayName="GrecoRelationships", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Greco Relationship Hydration Summary"
    ws_summary["A1"].font = Font(bold=True, size=14, color="1F4E5F")
    ws_summary["A3"] = "Total records"
    ws_summary["B3"] = len(rows)

    priority = Counter(row.get("priority_tier", "") or "Unspecified" for row in rows)
    confidence = Counter(row.get("data_confidence", "") or "Unspecified" for row in rows)
    warm = Counter(row.get("warm_path_needed", "") or "No/unspecified" for row in rows)

    ws_summary["A5"] = "Priority tier"
    ws_summary["D5"] = "Data confidence"
    ws_summary["G5"] = "Warm path needed"
    for cell in ("A5", "D5", "G5"):
        ws_summary[cell].font = Font(bold=True, color="FFFFFF")
        ws_summary[cell].fill = header_fill

    for offset, (key, val) in enumerate(priority.most_common(), start=6):
        ws_summary[f"A{offset}"] = key
        ws_summary[f"B{offset}"] = val
    for offset, (key, val) in enumerate(confidence.most_common(), start=6):
        ws_summary[f"D{offset}"] = key
        ws_summary[f"E{offset}"] = val
    for offset, (key, val) in enumerate(warm.most_common(), start=6):
        ws_summary[f"G{offset}"] = key
        ws_summary[f"H{offset}"] = val

    ws_summary["A14"] = "Operating note"
    ws_summary["A14"].font = Font(bold=True, color="1F4E5F")
    ws_summary["A15"] = (
        "This workbook is a relationship-intelligence seed, not an automated outreach list. "
        "Every Tier 1 route should be re-verified and context-briefed before contact."
    )
    ws_summary["A15"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_summary.merge_cells("A15:H17")

    for col in ["A", "D", "G"]:
        ws_summary.column_dimensions[col].width = 28
    for col in ["B", "E", "H"]:
        ws_summary.column_dimensions[col].width = 14

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")


if __name__ == "__main__":
    build()
